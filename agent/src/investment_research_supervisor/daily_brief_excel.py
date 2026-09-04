"""Excel appendix for persisted Investment Research Daily Brief payloads."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


_HEADERS = [
    "股票代码", "公司", "行业", "估值状态", "现价", "合理价值低", "合理价值中", "合理价值高",
    "相对中位值差距", "历史支撑低", "历史支撑高",
]


def _sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(_HEADERS)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        support = dict(row.get("historical_support") or {})
        worksheet.append([
            row.get("stock_code"),
            row.get("company_name"),
            row.get("industry_name"),
            row.get("valuation_label"),
            row.get("current_price"),
            row.get("fair_value_low"),
            row.get("fair_value_mid"),
            row.get("fair_value_high"),
            row.get("valuation_gap_percent"),
            support.get("low"),
            support.get("high"),
        ])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column in worksheet.iter_cols(min_row=2, min_col=5, max_col=11):
        for cell in column:
            cell.number_format = "0.00"
    widths = [16, 18, 14, 14, 12, 14, 14, 14, 16, 14, 14]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width


def _price_condition_sheet(workbook: Workbook, digest: dict[str, Any]) -> None:
    """「今日价格条件」附录：与卡片第一节同一数据源，纯投影。"""
    worksheet = workbook.create_sheet("今日价格条件")
    headers = ["代码", "公司", "范围", "价格条件", "现价", "落点", "主动作", "原因"]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in list(digest.get("lines") or []):
        worksheet.append([
            row.get("stock_code"),
            row.get("company_name"),
            "在范围内" if row.get("eligibility_status") == "IN_VALUE_SCOPE" else "不在范围内",
            row.get("effective_label"),
            row.get("current_price"),
            row.get("position_sentence"),
            row.get("primary_action_label"),
            row.get("reason_short") or "",
        ])
    if not list(digest.get("lines") or []):
        worksheet.append(["—", "—", "—", "今日无价格条件变化。", None, None, None, None])
    omitted = int(digest.get("omitted_count") or 0)
    if omitted > 0:
        worksheet.append(["—", "—", "—", f"另有 {omitted} 家见公司研究页，未列入日报。", None, None, None, None])
    worksheet.freeze_panes = "A2"
    widths = [16, 18, 12, 26, 12, 26, 22, 36]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width
    for column in worksheet.iter_cols(min_row=2, min_col=5, max_col=5):
        for cell in column:
            cell.number_format = "0.00"


def export_daily_brief_workbook(brief: dict[str, Any], output_path: str | Path) -> Path:
    """Export persisted Daily Brief rows; the price-condition sheet comes first."""
    payload = dict(brief.get("brief_payload") or {})
    digest = dict(brief.get("price_condition_digest") or {})
    if not digest:
        digest = dict(payload.get("price_condition_digest") or {})
    workbook = Workbook()
    workbook.remove(workbook.active)
    _price_condition_sheet(workbook, digest)
    _sheet(workbook, "低估龙头池", list(payload.get("low_value_leader_table") or []))
    _sheet(workbook, "深度低估", list(payload.get("deeply_undervalued_companies") or []))
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path
