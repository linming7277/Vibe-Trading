from __future__ import annotations

from pathlib import Path

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from src.api import research_supervisor_routes
from src.focus_selection import FocusSelectionService
from src.investment_research_supervisor.daily_brief_bitable_service import (
    DailyBriefBitablePublisher,
    MANAGED_SOURCE,
    _coerce_for_bitable_fields,
)
from src.investment_research_supervisor.daily_brief_excel import export_daily_brief_workbook
from src.investment_research_supervisor.daily_brief_table_image import render_value_observation_table
from src.financial_analysis.store import FinancialAnalysisStore
from src.investment_research_supervisor.daily_brief_notification_service import (
    DailyBriefNotificationService,
    DailyBriefNotificationSettings,
    build_daily_brief_card,
)
from src.investment_research_supervisor.hermes_feishu import HermesSupervisorFeishuCredentials
from src.investment_research_supervisor.daily_brief_service import InvestmentResearchDailyBriefService
from src.investment_research_supervisor.daily_brief_store import InvestmentResearchDailyBriefRepository
from src.low_value_leader_pool.store import LowValueLeaderPoolRepository
from src.low_value_risk_snapshot.store import LowValueRiskSnapshotRepository
from src.risk_research_preparation.store import RiskResearchPreparationRepository


AS_OF = "2026-08-25"


def test_hermes_supervisor_credentials_require_credentials_and_home_channel() -> None:
    credentials = HermesSupervisorFeishuCredentials.load(values={
        "FEISHU_APP_ID": "cli_supervisor",
        "FEISHU_APP_SECRET": "secret",
        "FEISHU_HOME_CHANNEL": "oc_daily_brief",
        "FEISHU_DOMAIN": "feishu",
    })

    assert credentials.app_id == "cli_supervisor"
    assert credentials.target_id == "oc_daily_brief"
    assert credentials.domain == "feishu"


def test_hermes_supervisor_credentials_reject_missing_home_channel() -> None:
    import pytest

    with pytest.raises(RuntimeError, match="home channel"):
        HermesSupervisorFeishuCredentials.load(values={
            "FEISHU_APP_ID": "cli_supervisor", "FEISHU_APP_SECRET": "secret",
        })


def test_bitable_payload_respects_existing_text_column_type() -> None:
    payload = _coerce_for_bitable_fields(
        {"相对中位值差距": 42.5, "现价": 10.2},
        {"相对中位值差距": 1, "现价": 2},
    )

    assert payload == {"相对中位值差距": "42.5", "现价": 10.2}


def _item(code: str, *, deeply_undervalued: bool = False) -> dict:
    return {
        "market": "CN",
        "stock_code": code,
        "company_name": f"公司{code[-2:]}",
        "industry_code": "I1",
        "industry_name": "示例行业",
        "leader_rank": 1,
        "leader_score": 90.0,
        "current_price": 10.0,
        "fair_value_low": 11.0,
        "fair_value_mid": 12.0,
        "fair_value_high": 13.0,
        "valuation_status": "DEEPLY_UNDERVALUED" if deeply_undervalued else "UNDERVALUED",
        "historical_valuation_status": "CHEAP",
        "support_status": "AVAILABLE",
        "support_zone_low": 9.5,
        "support_zone_high": 10.5,
        "entry_level": "WATCH",
        "source_pool_id": "pool-1",
        "source_as_of": AS_OF,
        "enter_reason": "PRICE_ENTERED_LOW_VALUE",
        "metadata": {
            "data_quality": {
                "price": {"status": "READY", "as_of": f"{AS_OF}T15:00:00+08:00"},
            },
        },
    }


class FakeEntryResearchService:
    def __init__(self, supports: dict[str, dict | None] | None = None) -> None:
        self.supports = supports or {}
        self.calls: list[tuple[str, str, str | None]] = []

    def get_entry_research(self, market: str, stock_code: str, *, as_of: str | None = None) -> dict:
        self.calls.append((market, stock_code, as_of))
        return {"focus_zones": {"historical_support": self.supports.get(stock_code)}}


class FakeFocusSelection:
    """Deterministic stand-in returning an explicit A-tier list."""

    def __init__(self, a_items: list[dict] | None = None) -> None:
        self.a_items = a_items or [{
            "stock_code": "000002.SZ",
            "focus_reasons": ["当前处于深度低估区域"],
            "focus_cautions": ["公司核心逻辑由 AI 初步形成，待人工复核"],
        }]
        self.calls: list[str | None] = []

    def get_focus_selection(self, *, as_of: str | None = None) -> dict:
        self.calls.append(as_of)
        return {"research_as_of": as_of, "A": self.a_items, "A_count": len(self.a_items)}


def _seed(
    tmp_path: Path,
    *,
    focus_selection_service=None,
) -> tuple[InvestmentResearchDailyBriefService, InvestmentResearchDailyBriefRepository]:
    db_path = tmp_path / "research.db"
    briefs = InvestmentResearchDailyBriefRepository(db_path)
    pool = LowValueLeaderPoolRepository(db_path)
    risk = LowValueRiskSnapshotRepository(db_path)
    financial = FinancialAnalysisStore(db_path)
    first = _item("000001.SZ")
    second = _item("000002.SZ", deeply_undervalued=True)
    pool.synchronize_refresh(
        eligible={first["stock_code"]: first, second["stock_code"]: second},
        current_codes={first["stock_code"], second["stock_code"]},
        evaluated={}, error_codes=set(), source_pool_id="pool-1", source_as_of=AS_OF,
        remove_reason=lambda _status: "VALUATION_RECOVERED",
    )
    risk.save({
        "market": "CN", "stock_code": first["stock_code"], "source_as_of": AS_OF,
        "overall_risk": "HIGH", "value_trap_risk": "HIGH", "material_risk_count": 1,
        "high_risk_count": 1, "medium_risk_count": 0, "top_risk_types": ["现金流"],
        "risk_summary": "现金流资料需要重点核验。", "financial_status": "READY",
        "business_status": "READY", "thesis_status": "READY", "formula_version": "risk-v1",
    })
    risk.save({
        "market": "CN", "stock_code": second["stock_code"], "source_as_of": AS_OF,
        "overall_risk": "MEDIUM", "value_trap_risk": "MEDIUM", "material_risk_count": 1,
        "high_risk_count": 0, "medium_risk_count": 1, "top_risk_types": ["行业"],
        "risk_summary": "行业资料需要复核。", "financial_status": "READY",
        "business_status": "READY", "thesis_status": "READY", "formula_version": "risk-v1",
    })
    financial.save_python_snapshot({
        "stock_code": first["stock_code"], "stock_name": first["company_name"], "as_of": AS_OF,
        "historical_cutoff": AS_OF, "financial_feature_version": "v1", "forecast_version": "v1",
        "feature_status": "READY", "forecast_status": "READY", "analysis_status": "COMPLETED",
        "identity": {}, "history": [],
        "feature": {"latest_changes": [{
            "metric": "revenue", "previous": 100, "current": 110, "change_percent": 10,
            "report_date": "2026-06-30", "fact": True,
        }]},
        "forecast": {}, "data_gaps": [], "source_hash": "financial-1",
    })
    entry_research = FakeEntryResearchService({
        second["stock_code"]: {"low": 9.0, "high": 10.0, "strength": "HIGH"},
    })
    class _FakeWatchpoints:
        def get_watchpoints(self, *args, **kwargs):
            return {"top_watchpoints": []}

    service = InvestmentResearchDailyBriefService(
        repository=briefs, pool_repository=pool, risk_repository=risk, financial_store=financial,
        entry_research_service=entry_research,
        focus_selection_service=focus_selection_service or FakeFocusSelection(),
        watchpoint_projection_service=_FakeWatchpoints(),
        web_base_url="https://research.example.test",
    )
    thesis = service.thesis_repository.create_initial_thesis({
        "market": "CN", "stock_code": second["stock_code"], "title": "示例 Thesis",
        "core_thesis": "已有核心逻辑", "status": "WEAKENING", "confidence": "MEDIUM",
        "invalid_conditions": [], "created_by": "HUMAN", "updated_by": "HUMAN",
        "source_data_as_of": AS_OF,
    })
    service.evidence_repository.create_evidence({
        "thesis_id": thesis["thesis_id"], "market": "CN", "stock_code": second["stock_code"],
        "evidence_type": "BUSINESS", "effect": "CHALLENGE", "claim": "需求变化",
        "summary": "已有资料显示需求变化。", "source_type": "COMPANY_RESEARCH_SNAPSHOT",
        "data_as_of": AS_OF, "confidence": "HIGH", "created_by": "SYSTEM",
    })
    return service, briefs


def test_daily_brief_uses_same_day_inputs_persists_once_and_preserves_priority_order(tmp_path: Path) -> None:
    service, repository = _seed(tmp_path)

    built = service.build(research_as_of=AS_OF)

    assert built.status == "READY"
    assert not built.reused
    brief = built.brief
    assert brief["research_as_of"] == AS_OF
    assert brief["low_value_active_count"] == 2
    assert brief["enter_count"] == 2
    assert brief["exit_count"] == 0
    assert brief["priority_companies"][0]["stock_code"] == "000001.SZ"
    assert brief["priority_companies"][1]["stock_code"] == "000002.SZ"
    assert brief["risk_summary"]["high_risk_review_count"] == 1
    assert brief["risk_summary"]["risk_review_count"] == 1
    assert brief["thesis_changes"] == [{
        "stock_code": "000002.SZ", "company_name": "公司SZ",
        "changes": ["Thesis WEAKENING", "新增 active Challenge Evidence"],
        "challenge_evidence": ["已有资料显示需求变化。"],
        "previous_status": None,
        "thesis_status": "WEAKENING", "source_data_as_of": AS_OF,
    }]
    assert brief["financial_changes"][0]["source_as_of"] == AS_OF
    assert brief["formula_version"] == "daily-brief-v25"
    deep_list = brief["brief_payload"]["deeply_undervalued_companies"]
    assert brief["brief_payload"]["deeply_undervalued_count"] == 1
    assert deep_list[0]["stock_code"] == "000002.SZ"
    assert deep_list[0]["valuation_gap_percent"] == 20.0
    assert deep_list[0]["historical_support"] == {
        "status": "READY", "low": 9.0, "high": 10.0, "strength": "HIGH",
    }
    assert "leader_score" not in deep_list[0]
    executive_text = brief["brief_payload"]["text"]
    assert "二、深度低估观察名单" not in executive_text
    assert "风险复核" not in executive_text
    assert "资料不足" not in executive_text
    assert "二、今日投资判断变化" in executive_text
    assert "三、重点研究观察" in executive_text
    assert "与「机会与风险」页 A 级重点研究一致" in executive_text
    assert "公司SZ / 000002.SZ" in executive_text
    assert "| 11.0 / 12.0 / 13.0 | 20.00% | 9.0–10.0 |" in executive_text
    assert "四、低估龙头表格" in executive_text
    assert "当前低估龙头池（不保留历史）" in executive_text
    assert "https://acnhfzsa8929.feishu.cn/base/WOxgbNUrVagmjCsfNNZcXCySndh" in executive_text
    assert brief["brief_payload"]["low_value_leader_bitable_url"].startswith("https://acnhfzsa8929.feishu.cn/base/")
    assert "revenue" not in executive_text
    assert "买入" not in executive_text
    watchlist = brief["brief_payload"]["executive_watchlist"]
    assert len(watchlist) == 1
    assert watchlist[0]["stock_code"] == "000002.SZ"
    assert watchlist[0]["current_price"] == 10.0
    assert watchlist[0]["fair_value_low"] == 11.0
    assert watchlist[0]["fair_value_mid"] == 12.0
    assert watchlist[0]["fair_value_high"] == 13.0
    assert watchlist[0]["historical_support"] == {
        "status": "READY", "low": 9.0, "high": 10.0, "strength": "HIGH",
    }
    assert watchlist[0]["research_change"] == "核心研究逻辑转弱；出现新增挑战证据"
    assert watchlist[0]["research_priority_reason"] == "当前处于深度低估区域"
    assert watchlist[0]["research_cautions"] == ["公司核心逻辑由 AI 初步形成，待人工复核"]
    assert watchlist[0]["risk_summary"] == "行业资料需要复核。"
    assert watchlist[0]["valuation_caveat"] is None
    assert brief["brief_payload"]["executive_watchlist_basis"] == "FOCUS_A"
    assert service._valuation_caveat(300) == "现价与估值中枢偏离显著，需要优先核验盈利与估值假设"
    assert brief["brief_payload"]["research_appendix"]["risk_summary"]["high_risk_review_count"] == 1
    assert repository.get_completed(AS_OF)["id"] == brief["id"]

    reused = service.build(research_as_of=AS_OF)
    assert reused.reused
    assert reused.brief["id"] == brief["id"]


def test_daily_brief_card_renders_value_observations_as_readable_summaries(tmp_path: Path) -> None:
    service, _ = _seed(tmp_path)
    brief = service.build(research_as_of=AS_OF).brief

    card = build_daily_brief_card(brief)
    content = "\n".join(str(item.get("content") or "") for item in card["elements"])

    assert card["header"]["title"]["content"] == "投研主管｜每日简报"
    assert "**研究日期** 2026-08-25" in content
    assert "**重点研究 · 1 家**" in content
    assert "**1. 公司SZ / 000002.SZ**" in content
    assert "合理价值中枢 **12.0**" in content
    assert "合理价值范围 11.0–13.0　｜　历史支撑 9.0–10.0" in content
    assert not any(item["tag"] in {"column_set", "img"} for item in card["elements"])
    assert any(item["tag"] == "action" for item in card["elements"])


def test_daily_brief_card_keeps_responsive_rows_when_image_key_is_supplied(tmp_path: Path) -> None:
    service, _ = _seed(tmp_path)
    brief = service.build(research_as_of=AS_OF).brief
    output = render_value_observation_table(brief, tmp_path / "重点研究观察.png")

    assert output.exists()
    assert output.stat().st_size > 1_000
    card = build_daily_brief_card(brief, value_table_image_key="img_value_table")
    assert not any(item["tag"] == "img" for item in card["elements"])
    assert any("公司SZ / 000002.SZ" in str(item.get("content") or "") for item in card["elements"])


def test_daily_brief_dry_run_is_idempotent_and_does_not_call_sender(tmp_path: Path) -> None:
    service, repository = _seed(tmp_path)
    service.build(research_as_of=AS_OF)

    class Sender:
        def send_interactive_card(self, **_kwargs):
            raise AssertionError("dry run must not send")

    notifier = DailyBriefNotificationService(
        repository=repository,
        settings=DailyBriefNotificationSettings(enabled=True, target_id="oc_test", dry_run=True),
        sender=Sender(),
    )

    first = notifier.notify(research_as_of=AS_OF)
    second = notifier.notify(research_as_of=AS_OF)

    assert first["status"] == "READY"
    assert first["covers_low_value"] is True
    assert "投研主管｜每日简报" == first["card"]["header"]["title"]["content"]
    assert "研究日期" in first["card"]["elements"][0]["content"]
    assert not any(element["tag"] == "column_set" for element in first["card"]["elements"])
    # Dry-run has no published Bitable delivery, so it intentionally omits
    # the final navigation button.
    assert not any(element.get("tag") == "action" for element in first["card"]["elements"])
    assert second["status"] == "REUSED"
    assert second["covers_low_value"] is True


def test_daily_brief_api_reads_completed_record_without_building(tmp_path: Path, monkeypatch) -> None:
    service, _ = _seed(tmp_path)
    service.build(research_as_of=AS_OF)
    app = FastAPI()

    async def allow():
        return None

    monkeypatch.setattr(research_supervisor_routes, "get_investment_research_daily_brief_service", lambda: service)
    research_supervisor_routes.register_research_supervisor_routes(app, allow)
    client = TestClient(app)

    response = client.get(f"/api/research-supervisor/daily-brief?as_of={AS_OF}")
    missing = client.get("/api/research-supervisor/daily-brief?as_of=2026-08-24")

    assert response.status_code == 200
    assert response.json()["research_as_of"] == AS_OF
    assert missing.status_code == 404


def test_daily_brief_displays_at_most_ten_focus_a_companies(tmp_path: Path) -> None:
    service, _ = _seed(tmp_path)
    companies = [_item("000001.SZ"), _item("000002.SZ", deeply_undervalued=True)]
    for index in range(3, 14):
        company = _item(f"{index:06d}.SZ", deeply_undervalued=True)
        company["leader_score"] = 100.0 - index
        if index == 11:
            company["current_price"] = None
        companies.append(company)
    eligible = {company["stock_code"]: company for company in companies}
    service.pool_repository.synchronize_refresh(
        eligible=eligible,
        current_codes=set(eligible),
        evaluated={},
        error_codes=set(),
        source_pool_id="pool-1",
        source_as_of=AS_OF,
        remove_reason=lambda _status: "VALUATION_RECOVERED",
    )
    service.focus_selection_service = FakeFocusSelection([
        {"stock_code": company["stock_code"], "focus_reasons": ["当前处于低估区域"], "focus_cautions": []}
        for company in companies[2:]
    ])

    brief = service.build(research_as_of=AS_OF).brief
    payload = brief["brief_payload"]
    deep_list = payload["deeply_undervalued_companies"]

    assert payload["deeply_undervalued_count"] == 12
    assert len(deep_list) == 12
    assert [item["stock_code"] for item in deep_list[:3]] == ["000003.SZ", "000004.SZ", "000005.SZ"]
    assert deep_list[9]["stock_code"] == "000011.SZ"
    assert deep_list[9]["valuation_gap_percent"] is None
    assert "Leader Score" not in payload["text"]
    assert "000013.SZ｜" not in payload["text"]
    assert "深度低估观察名单" not in payload["text"]
    assert len(payload["executive_situations"]) <= 3
    assert len(payload["executive_watchlist"]) == 10
    assert "000011.SZ" not in [item["stock_code"] for item in payload["executive_watchlist"]]
    assert len(payload["low_value_leader_table"]) == len(companies)
    assert all("leader_score" not in item for item in payload["executive_watchlist"])
    assert all("leader_score" not in item for item in payload["low_value_leader_table"])


def test_daily_brief_watchlist_follows_real_focus_selection_a_tier(tmp_path: Path) -> None:
    """重点观察必须等于机会与风险的 A 级名单，而不是低估池的深度低估切片。"""
    service, _ = _seed(tmp_path)
    service.focus_selection_service = FocusSelectionService(
        pool_repository=service.pool_repository,
        risk_snapshot_repository=service.risk_repository,
        thesis_repository=service.thesis_repository,
    )
    eligible = {
        "000001.SZ": _item("000001.SZ"),
        "000002.SZ": _item("000002.SZ", deeply_undervalued=True),
        "000003.SZ": _item("000003.SZ"),
    }
    service.pool_repository.synchronize_refresh(
        eligible=eligible, current_codes=set(eligible), evaluated={}, error_codes=set(),
        source_pool_id="pool-1", source_as_of=AS_OF,
        remove_reason=lambda _status: "VALUATION_RECOVERED",
    )
    # 000003：风险低、资料就绪、核心逻辑有效，是唯一能进 A 级的公司；
    # 它并非深度低估，旧逻辑下根本不会出现在重点观察里。
    service.risk_repository.save({
        "market": "CN", "stock_code": "000003.SZ", "source_as_of": AS_OF,
        "overall_risk": "LOW", "value_trap_risk": "LOW", "material_risk_count": 0,
        "high_risk_count": 0, "medium_risk_count": 0, "top_risk_types": [],
        "risk_summary": "", "financial_status": "READY",
        "business_status": "READY", "thesis_status": "READY", "formula_version": "risk-v1",
    })
    RiskResearchPreparationRepository(service.pool_repository.db_path).upsert({
        "market": "CN", "stock_code": "000003.SZ", "research_as_of": AS_OF,
        "company_name": "公司SZ", "financial_status": "READY", "business_profile_status": "READY",
        "overall_status": "READY",
    })
    service.thesis_repository.create_initial_thesis({
        "market": "CN", "stock_code": "000003.SZ", "title": "示例 Thesis",
        "core_thesis": "已有核心逻辑", "status": "UNCHANGED", "confidence": "MEDIUM",
        "invalid_conditions": [], "created_by": "HUMAN", "updated_by": "HUMAN",
        "source_data_as_of": AS_OF,
    })
    thesis_conn = service.thesis_repository._conn
    thesis_conn.execute(
        "UPDATE company_theses SET created_at='2026-08-20T10:00:00', updated_at='2026-08-20T10:00:00' "
        "WHERE market='CN' AND stock_code='000003.SZ'",
    )
    thesis_conn.commit()

    brief = service.build(research_as_of=AS_OF).brief
    payload = brief["brief_payload"]
    watchlist = payload["executive_watchlist"]

    # 000001 高风险被硬降 C；000002 深度低估但当日无有效核心逻辑被软降 B。
    assert [item["stock_code"] for item in watchlist] == ["000003.SZ"]
    assert payload["executive_watchlist_basis"] == "FOCUS_A"
    assert watchlist[0]["research_priority_reason"].startswith("当前处于低估区域")
    assert "leader_score" not in watchlist[0]
    assert "与「机会与风险」页 A 级重点研究一致" in payload["text"]


def test_daily_brief_watchlist_falls_back_to_deep_list_when_focus_selection_unavailable(
    tmp_path: Path,
) -> None:
    class BrokenFocusSelection:
        def get_focus_selection(self, *, as_of: str | None = None) -> dict:
            raise RuntimeError("FOCUS_SELECTION_AS_OF_UNAVAILABLE")

    service, _ = _seed(tmp_path, focus_selection_service=BrokenFocusSelection())

    brief = service.build(research_as_of=AS_OF).brief
    payload = brief["brief_payload"]

    assert payload["executive_watchlist_basis"] == "DEEP_FALLBACK"
    assert [item["stock_code"] for item in payload["executive_watchlist"]] == ["000002.SZ"]
    assert payload["executive_watchlist"][0]["research_priority_reason"] == "当前处于深度低估状态"
    assert any("机会与风险筛选不可用" in gap for gap in brief["data_gaps"])
    assert "沿用深度低估名单" in payload["text"]


def test_daily_brief_investment_judgment_changes_require_a_real_delta(tmp_path: Path) -> None:
    service, _ = _seed(tmp_path)
    current = [{
        "stock_code": "000001.SZ", "company_name": "公司SZ", "valuation_label": "深度低估",
        "fair_value_mid": 12.0,
    }]
    previous = {
        "research_as_of": "2026-08-24",
        "brief_payload": {"low_value_leader_table": [{
            "stock_code": "000001.SZ", "company_name": "公司SZ", "valuation_label": "低估",
            "fair_value_mid": 10.0,
        }]},
    }

    situations = service._executive_situations(
        risks={}, thesis_changes=[], low_value_leader_table=current, previous_brief=previous,
    )
    no_delta = service._executive_situations(
        risks={}, thesis_changes=[], low_value_leader_table=current, previous_brief=None,
    )

    assert situations == [{
        "stock_code": "000001.SZ", "company_name": "公司SZ",
        "basis": "合理价值中枢较 2026-08-24 上调 20.00%",
        "impact": "估值锚已实质上移，需要重新评估上行空间与盈利兑现条件。",
    }]
    assert no_delta == []


def test_daily_brief_rebuilds_ready_record_when_template_version_changes(tmp_path: Path) -> None:
    service, repository = _seed(tmp_path)
    previous_payload = service._build_payload(AS_OF)
    previous_payload["formula_version"] = "daily-brief-v7"
    repository.save_ready(previous_payload)

    rebuilt = service.build(research_as_of=AS_OF)
    reused = service.build(research_as_of=AS_OF)

    assert rebuilt.status == "READY"
    assert not rebuilt.reused
    assert rebuilt.brief["formula_version"] == "daily-brief-v25"
    assert "deeply_undervalued_companies" in rebuilt.brief["brief_payload"]
    assert reused.reused


def test_daily_brief_deep_list_uses_pit_historical_support_and_valuation_range(tmp_path: Path) -> None:
    service, _ = _seed(tmp_path)
    entry_research = service.entry_research_service

    brief = service.build(research_as_of=AS_OF).brief
    item = brief["brief_payload"]["deeply_undervalued_companies"][0]

    assert item["fair_value_low"] == 11.0
    assert item["fair_value_mid"] == 12.0
    assert item["fair_value_high"] == 13.0
    assert item["historical_support"] == {
        "status": "READY", "low": 9.0, "high": 10.0, "strength": "HIGH",
    }
    assert entry_research.calls == [("CN", "000002.SZ", AS_OF)]
    assert "deeply_undervalued_observation_count" not in item
    assert "first_deeply_undervalued_as_of" not in item
    assert "previous_observation" not in item
    assert "估值区间（低/中/高）" not in brief["brief_payload"]["text"]
    assert brief["brief_payload"]["research_appendix"]["data_gaps"]


def test_daily_brief_marks_missing_historical_support_as_insufficient(tmp_path: Path) -> None:
    service, _ = _seed(tmp_path)
    service.entry_research_service = FakeEntryResearchService()

    brief = service.build(research_as_of=AS_OF).brief
    item = brief["brief_payload"]["deeply_undervalued_companies"][0]

    assert item["historical_support"] == {
        "status": "INSUFFICIENT", "low": None, "high": None, "strength": None,
    }
    assert "资料不足" not in brief["brief_payload"]["text"]


def test_daily_brief_exports_low_value_leader_excel(tmp_path: Path) -> None:
    service, _ = _seed(tmp_path)
    brief = service.build(research_as_of=AS_OF).brief
    output = export_daily_brief_workbook(brief, tmp_path / "低估龙头表格.xlsx")

    workbook = load_workbook(output, data_only=True)

    assert workbook.sheetnames == ["低估龙头池", "深度低估"]
    assert [cell.value for cell in workbook["低估龙头池"][1]] == [
        "股票代码", "公司", "行业", "估值状态", "现价", "合理价值低", "合理价值中", "合理价值高",
        "相对中位值差距", "历史支撑低", "历史支撑高",
    ]
    assert workbook["低估龙头池"].max_row == 3
    assert workbook["深度低估"].max_row == 2
    # 低估龙头池 sheet 的支撑列来自池快照区间，不再因为未现算而留空。
    assert workbook["低估龙头池"].cell(row=2, column=10).value == 9.5
    assert workbook["低估龙头池"].cell(row=2, column=11).value == 10.5
    assert "Leader Score" not in [cell.value for cell in workbook["低估龙头池"][1]]


def test_daily_brief_retries_only_missing_card_after_attachment(tmp_path: Path) -> None:
    service, repository = _seed(tmp_path)
    service.build(research_as_of=AS_OF)

    class Sender:
        def __init__(self) -> None:
            self.calls: list[str] = []
            self.fail_card_once = True

        def send_file(self, **_kwargs) -> str:
            self.calls.append("file")
            return "file-message"

        def send_interactive_card(self, **_kwargs) -> str:
            self.calls.append("card")
            if self.fail_card_once:
                self.fail_card_once = False
                raise RuntimeError("card unavailable")
            return "card-message"

    sender = Sender()
    notifier = DailyBriefNotificationService(
        repository=repository,
        settings=DailyBriefNotificationSettings(enabled=True, target_id="oc_test", dry_run=False),
        sender=sender,
    )

    first = notifier.notify(research_as_of=AS_OF)
    second = notifier.notify(research_as_of=AS_OF)
    third = notifier.notify(research_as_of=AS_OF)

    assert first["status"] == "FAILED"
    assert first["delivery"]["attachment_message_id"] == "file-message"
    assert second["status"] == "READY"
    assert second["delivery"]["message_id"] == "card-message"
    assert second["delivery"]["attachment_message_id"] == "file-message"
    assert third["status"] == "REUSED"
    assert sender.calls == ["file", "card", "card"]


def test_bitable_publisher_syncs_current_pool_and_keeps_manual_rows(tmp_path: Path) -> None:
    service, repository = _seed(tmp_path)
    brief = service.build(research_as_of=AS_OF).brief

    class Gateway:
        def __init__(self) -> None:
            self.fields = [
                {"field_id": "fld_primary", "field_name": "文本"},
                {"field_id": "fld_fair_low", "field_name": "合理价值低"},
                {"field_id": "fld_fair_mid", "field_name": "合理价值中"},
                {"field_id": "fld_fair_high", "field_name": "合理价值高"},
                {"field_id": "fld_support_low", "field_name": "历史支撑低"},
                {"field_id": "fld_support_high", "field_name": "历史支撑高"},
            ]
            self.records = [
                {"record_id": "managed-current", "fields": {
                    "同步来源": MANAGED_SOURCE, "股票代码": "000001.SZ", "研究日期": AS_OF,
                }},
                {"record_id": "legacy-current", "fields": {
                    "股票代码": "000002.SZ", "研究日期": AS_OF, "日报版本": "daily-brief-v11",
                }},
                {"record_id": "managed-stale", "fields": {
                    "同步来源": MANAGED_SOURCE, "股票代码": "999999.SZ", "研究日期": "2026-08-24",
                }},
                {"record_id": "legacy-stale", "fields": {
                    "文本": "2026-08-24|888888.SZ", "股票代码": "888888.SZ",
                    "研究日期": "2026-08-24", "日报版本": "daily-brief-v15",
                }},
                {"record_id": "manual", "fields": {"股票代码": "MANUAL.SZ", "研究日期": "2026-08-24"}},
            ]
            self.created: list[dict] = []
            self.updated: list[dict] = []
            self.deleted: list[str] = []
            self.deleted_fields: list[str] = []

        def list_fields(self) -> list[dict]:
            return self.fields

        def create_text_field(self, name: str) -> None:
            self.fields.append({"field_name": name})

        def delete_field(self, field_id: str) -> None:
            self.deleted_fields.append(field_id)

        def list_records(self) -> list[dict]:
            return self.records

        def batch_create(self, records: list[dict]) -> None:
            self.created.extend(records)

        def batch_update(self, records: list[dict]) -> None:
            self.updated.extend(records)

        def batch_delete(self, record_ids: list[str]) -> None:
            self.deleted.extend(record_ids)

    gateway = Gateway()
    result = DailyBriefBitablePublisher(repository=repository, gateway=gateway).publish(research_as_of=AS_OF)

    assert result["status"] == "READY"
    assert result["row_count"] == 2
    assert result["created"] == 0
    assert result["updated"] == 2
    assert result["deleted"] == 2
    assert gateway.deleted == ["managed-stale", "legacy-stale"]
    assert {row["record_id"] for row in gateway.updated} == {"managed-current", "legacy-current"}
    assert all(row["fields"]["同步来源"] == MANAGED_SOURCE for row in gateway.updated)
    assert all("Leader Score" not in row["fields"] for row in gateway.updated)
    assert all("合理价值范围" in row["fields"] for row in gateway.updated)
    assert all("历史支撑范围" in row["fields"] for row in gateway.updated)
    support_by_code = {row["fields"]["股票代码"]: row["fields"]["历史支撑范围"] for row in gateway.updated}
    assert support_by_code["000001.SZ"] == "9.5–10.5元"  # 非深度低估/A级：回退池快照支撑区间
    assert support_by_code["000002.SZ"] == "9–10元"      # 深度低估/A级：优先简报现算支撑
    assert all("合理价值低" not in row["fields"] for row in gateway.updated)
    assert result["retired_fields"] == 5
    assert gateway.deleted_fields == ["fld_fair_low", "fld_fair_mid", "fld_fair_high", "fld_support_low", "fld_support_high"]
    assert "同步来源" in {item["field_name"] for item in gateway.fields}
    assert repository.delivery(
        research_as_of=AS_OF, channel="feishu_bitable", target_id="tblJb3Pc7w9fKsjI",
    )["status"] == "SENT"
    assert brief["brief_payload"]["low_value_leader_table"]


def test_daily_brief_uses_bitable_link_without_excel_after_publication(tmp_path: Path) -> None:
    service, repository = _seed(tmp_path)
    service.build(research_as_of=AS_OF)
    repository.record_delivery(
        research_as_of=AS_OF, channel="feishu_bitable", target_id="tblJb3Pc7w9fKsjI",
        status="SENT", message_id="https://acnhfzsa8929.feishu.cn/base/test",
    )

    class Sender:
        def __init__(self) -> None:
            self.calls: list[str] = []

        def send_file(self, **_kwargs) -> str:
            self.calls.append("file")
            raise AssertionError("published Bitable must replace the Excel attachment")

        def upload_image(self, *, file_path: str) -> str:
            self.calls.append("table")
            assert Path(file_path).suffix == ".png"
            assert Path(file_path).exists()
            return "img_value_table"

        def send_interactive_card(self, **_kwargs) -> str:
            self.calls.append("card")
            self.card = _kwargs["card"]
            return "card-message"

    sender = Sender()
    notifier = DailyBriefNotificationService(
        repository=repository,
        settings=DailyBriefNotificationSettings(enabled=True, target_id="oc_test", dry_run=False),
        sender=sender,
    )

    result = notifier.notify(research_as_of=AS_OF)

    assert result["status"] == "READY"
    assert sender.calls == ["card"]
    assert not any(item["tag"] == "img" for item in sender.card["elements"])
    assert result["delivery"]["attachment_message_id"] is None
